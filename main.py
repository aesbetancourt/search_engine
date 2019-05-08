# When a wrote this, only God and I understood what I was doing. Now, God only knows.
__author__ = "Alejandro Sanchez"
__copyright__ = "Copyright 2019, Structure Search Engine"
__credits__ = ["Alejandro Sanchez, Gian Paolo Falconi"]
__license__ = "GPL"
__version__ = "1.0.0"
__maintainer__ = "Alejandro Sanchez"
__email__ = "aesbetancourt@outlook.com"
__status__ = "Prototype"

from functools import partial
import excel2img
from PyQt5 import QtWidgets, uic, QtCore, QtGui
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QTableWidgetItem, QMessageBox, QPushButton
from openpyxl import load_workbook
from PIL import Image

# Global variable and lists
# Filter 1
ranges = list()
year = list()
tipologia = list()
installation = list()
turbina = list()
structure = list()
exposition = list()

# Filter 2
seismicloads = list()
windloads = list()
blastloads = list()
thermalloads = list()
snowloads = list()
transport = list()

# Filter 3
definit = list()

# Others
module = ''
flag = 0
projectname = list()


# General Functions
def show_question(title, message):
    box = QMessageBox.question(None, title, message, QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
    if box == QMessageBox.Yes:
        return 1
    else:
        return 0


def show_message(title, message):
    QMessageBox.information(None, title, message)


def next_modules(previous=None, following=None):
    global module
    if module1.enclosure.currentText() == 'ON-BASE' and previous is not None and following is None:
        module2.show()
        previous.close()
        module = 'ON-BASE'
    elif module1.enclosure.currentText() == 'OFF-BASE' and previous is not None and following is None:
        module3.show()
        previous.close()
        module = 'OFF-BASE'
    elif previous is not None and following is not None:
        previous.close()
        following.show()
    elif following is not None:
        module2.close()
        module3.close()
        following.show()


def out():
    ans = show_question("Confirm Exit", "Are you sure you want to exit?")
    if ans == 1:
        try:
            # Delete all .png generated
            import os, glob
            os.chdir('modules/ranges')
            for file_name in glob.glob('*.png'):
                os.remove(file_name)
            quit()
        except:
            quit()


def back():
    global flag
    if flag == 0:
        next_modules(previous=module4, following=module1)
    elif flag == 1:
        flag = 0
        if module == 'ON-BASE':
            next_modules(previous=module2, following=module1)
            module2.wind.setChecked(False)
            module2.seismic.setChecked(False)
            module2.snow.setChecked(False)
            module2.transport.setChecked(False)
            module2.termical.setChecked(False)
            module2.blasting.setChecked(False)
        elif module == 'OFF-BASE':
            next_modules(previous=module3, following=module1)
            module3.wind.setChecked(False)
            module3.seismic.setChecked(False)
            module3.snow.setChecked(False)
            module3.transport.setChecked(False)
            module3.termical.setChecked(False)
            module3.blasting.setChecked(False)
    elif flag == 2:
        flag = 1
        if module == 'ON-BASE':
            next_modules(previous=module4, following=module2)
        elif module == 'OFF-BASE':
            next_modules(previous=module4, following=module3)
    # Clear table in module4 if return of see All
    while module4.table.rowCount() > 0:
        module4.table.removeRow(0)
    del definit[:]


# Functions in module1
def see_all():
    module4.table.setRowCount(len(tipologia))
    btns = tipologia.copy()
    chkBoxItem = tipologia.copy()
    for i in range(len(btns)):
        btns[i] = QPushButton(module4.table)
        btns[i].setText('VIEW')
        chkBoxItem[i] = QTableWidgetItem()
        chkBoxItem[i].setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
        chkBoxItem[i].setCheckState(QtCore.Qt.Unchecked)

    for row in range(len(tipologia)):
        module4.table.setItem(row, 0, QTableWidgetItem(projectname[row]))
        module4.table.setItem(row, 1, QTableWidgetItem(year[row]))
        module4.table.setItem(row, 2, QTableWidgetItem(turbina[row]))
        module4.table.setCellWidget(row, 3, btns[row])
        btns[row].clicked.connect(partial(show_image, row))
        module4.table.setItem(row, 4, chkBoxItem[row])
        row += row
    next_modules(module1, module4)


def next_advanced():
    global flag
    flag = 1
    next_modules(previous=module1)


def next_basic():
    turbine_name = module1.turbine_2.currentText()
    basic = list()
    basic_index = list()
    turbine_index = list()
    basic_filter = list()
    wind = module1.radio_wind.isChecked()
    seism = module1.radio_seismic.isChecked()
    snow = module1.radio_snow.isChecked()
    trans = module1.radio_transport.isChecked()
    thermal = module1.radio_thermal.isChecked()
    blasting = module1.radio_blasting.isChecked()

    if wind:
        basic = windloads.copy()
    elif seism:
        basic = seismicloads.copy()
    elif snow:
        basic = snowloads.copy()
    elif trans:
        basic = transport.copy()
    elif thermal:
        basic = thermalloads.copy()
    elif blasting:
        basic = blastloads.copy()
    else:
        print('not selected')

    for item in range(len(basic)):
        if basic[item] == 'SI' or basic[item] == 'SNOW LOAD:' or basic[item] == 'THERMAL':
            basic_index.append(item)
    #print(basic)
    for item in range(len(basic)):
        if turbina[item] == turbine_name:
            turbine_index.append(item)
    #print(turbine_index, basic_index)

    for item in basic_index:
        if (item not in basic_filter) and (item in turbine_index):
            basic_filter.append(item)
    #print(basic_filter)

    module4.table.setRowCount(len(basic_filter))
    btns = basic_filter.copy()
    chkBoxItem = basic_filter.copy()
    for i in range(len(btns)):
        btns[i] = QPushButton(module4.table)
        btns[i].setText('VIEW')
        chkBoxItem[i] = QTableWidgetItem()
        chkBoxItem[i].setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
        chkBoxItem[i].setCheckState(QtCore.Qt.Unchecked)

    for row in range(len(basic_filter)):
        module4.table.setItem(row, 0, QTableWidgetItem(projectname[basic_filter[row]]))
        module4.table.setItem(row, 1, QTableWidgetItem(year[basic_filter[row]]))
        module4.table.setItem(row, 2, QTableWidgetItem(turbina[basic_filter[row]]))
        module4.table.setCellWidget(row, 3, btns[row])
        module4.table.setItem(row, 4, chkBoxItem[row])
        btns[row].clicked.connect(partial(show_image, basic_filter[row]))
        row += row

    if len(basic_filter) < 1:
        show_message('Warning', "This structural verification, isn't present on this record, the calculations"
                                " must be done, please let it know to the person in charge of the calculations.")
    else:
        next_modules(previous=module1, following=module4)


def indices():
    # Lists of indices
    indextipo = list()
    indexinstall = list()
    indexturbine = list()
    indexconfig = list()
    indexexposition = list()
    # Get str values in module 1
    enclosure = module1.enclosure.currentText()
    install = module1.install.currentText()
    turbine = module1.turbine.currentText()
    configuration = module1.configuration.currentText()
    expo = module1.expo.currentText()
    # Filtered indexes that meet the condition
    for item in range(len(tipologia)):
        if tipologia[item] == enclosure:
            indextipo.append(item)
        if installation[item] == install:
            indexinstall.append(item)
        if turbina[item] == turbine:
            indexturbine.append(item)
        if structure[item] == configuration:
            indexconfig.append(item)
        if exposition[item] == expo:
            indexexposition.append(item)
    repeated = list()
    for item in indextipo:
        if (item not in repeated) and (item in indexinstall) and (item in indexturbine) and (item in indexconfig) and \
                (item in indexexposition):
            repeated.append(item)
    return repeated


# Functions in "about"
def about():
    QMessageBox.about(None, "Search Structure Engine", "<b>ABOUT THIS SOFTWARE:</b>\n"
                      "<p>This program consists of simplifying the modeling for the development of the project,"
                      " providing a standard pre-dimensioned of the structure with the characteristics specified"
                      " by the user.</p>"
                      "<p>It is advisable to have a specifications of the project to perform a more specific search.</p>"
                      "<b>METHOD OF PROCEDURE</b><p>The structure search engine functional as follows:</p>"
                      "<p>• Enter the data corresponding to the type of structure that the user wants to find,"
                      " if its not possession of such information the program allows to see"
                      " the record from the year 2010  until the current year.</p>"
                      "<p>• Once introduced the characteristics of location, type, model, etc."
                      " Then will request the charges applied in the enclosure.</p>"
                      "<p>• This will carry out a filtering with the entered specifications,"
                      " showing on the screen the result of the projects with the data introduced by"
                      " year, model, name of the project, etc.</p>"
                      "<p>• The user can access to the data sheet to see the structural configuration"
                      " presented in each of the filtered projects, in order to carry out a model with a criterion"
                      " of the most critical applied loads, or loads closest to those indicated in the project under"
                      " development.</p><p>Allowing again perform new searches.</p>"
                      "<p><b>Copyright © 2019, Structure Search Engine. All Rights Reserved.</b></p>")


# Functions in module2
def on_next():
    global flag
    flag = 2
    # Check if check checkboxes are selected
    if module == 'ON-BASE':
        wind = module2.wind.isChecked()
        seism = module2.seismic.isChecked()
        sno = module2.snow.isChecked()
        trans = module2.transport.isChecked()
        thermal = module2.termical.isChecked()
        blast = module2.blasting.isChecked()
    elif module == 'OFF-BASE':
        wind = module3.wind.isChecked()
        seism = module3.seismic.isChecked()
        sno = module3.snow.isChecked()
        trans = module3.transport.isChecked()
        thermal = module3.termical.isChecked()
        blast = module3.blasting.isChecked()

    indexseism = list()
    indexwind = list()
    indexblast = list()
    indexthermal = list()
    indexsnow = list()
    indextransport = list()

    # Lists of filters
    filtered = indices()
    repeated = list()
    lists = list()

    for item in range(len(tipologia)):
        try:
            if seismicloads[item] == 'SI':
                indexseism.append(item)
            if windloads[item] == 'SI':
                indexwind.append(item)
            if thermalloads[item] == 'THERMAL':
                indexthermal.append(item)
            if blastloads[item] == 'SI':
                indexblast.append(item)
            if snowloads[item] == 'SNOW LOAD:':
                indexsnow.append(item)
            if transport[item] == 'SI':
                indextransport.append(item)
        except:
            show_message('Writer Error', 'Please check the syntax of the fields on the .xslx file,'
                                         ' the results shown can be wrong. (3)')

    try:
        if not seism:
            # print("1-", indexseism)
            for item in indexseism:
                try:
                    indexwind.remove(item)
                except: pass
                try:
                    indexblast.remove(item)
                except: pass
                try:
                    indexthermal.remove(item)
                except: pass
                try:
                    indexsnow.remove(item)
                except: pass
                try:
                    indextransport.remove(item)
                except: pass
            indexseism = []
            # print("1-", indexseism)
        if not wind:
            for item in indexwind:
                try:
                    indexseism.remove(item)
                except: pass
                try:
                    indexblast.remove(item)
                except: pass
                try:
                    indexthermal.remove(item)
                except: pass
                try:
                    indexsnow.remove(item)
                except: pass
                try:
                    indextransport.remove(item)
                except: pass
            indexwind = []
        if not sno:
            for item in indexsnow:
                try:
                    indexwind.remove(item)
                except: pass
                try:
                    indexblast.remove(item)
                except: pass
                try:
                    indexthermal.remove(item)
                except: pass
                try:
                    indexseism.remove(item)
                except: pass
                try:
                    indextransport.remove(item)
                except: pass
            indexsnow = []
        if not trans:
            for item in indextransport:
                try:
                    indexwind.remove(item)
                except: pass
                try:
                    indexblast.remove(item)
                except: pass
                try:
                    indexthermal.remove(item)
                except: pass
                try:
                    indexsnow.remove(item)
                except: pass
                try:
                    indexseism.remove(item)
                except: pass
            indextransport = []
        if not thermal:
            for item in indexthermal:
                try:
                    indexwind.remove(item)
                except: pass
                try:
                    indexblast.remove(item)
                except: pass
                try:
                    indexseism.remove(item)
                except: pass
                try:
                    indexsnow.remove(item)
                except: pass
                try:
                    indextransport.remove(item)
                except: pass
            indexthermal = []
        if not blast:
            for item in indexblast:
                try:
                    indexwind.remove(item)
                except: pass
                try:
                    indexseism.remove(item)
                except: pass
                try:
                    indexthermal.remove(item)
                except: pass
                try:
                    indexsnow.remove(item)
                except: pass
                try:
                    indextransport.remove(item)
                except: pass
            indexblast = []
    except:
        show_message('FAIL', 'COD: 2')
    try:
        lists.append(indexseism)
        lists.append(indexwind)
        lists.append(indexthermal)
        lists.append(indexblast)
        lists.append(indexsnow)
        lists.append(indextransport)
        if thermal and len(indexthermal) == 0:
            lists.clear()
        elif sno and len(indexsnow) == 0:
            lists.clear()
        elif trans and len(indextransport) == 0:
            lists.clear()
        elif wind and len(indexwind) == 0:
            lists.clear()
        elif seism and len(indexseism) == 0:
            lists.clear()
        elif blast and len(indexblast) == 0:
            lists.clear()
        try:
            lists = [element for element in lists if element]
            if len(lists) == 1:
                repeated = lists[0]
            else:
                repeated = set(lists[0])

                for s in lists[1:]:
                    repeated.intersection_update(s)
        except:
            pass

        # Only for debug
        #print('Listas de listas:', lists)
        #print('Filtro 1: ', filtered)
        #print('Filtro 2: ', repeated)

        for item in filtered:
            if (item not in definit) and (item in repeated):
                definit.append(item)
        #print('Filtro 3:', definit)
    except:
        show_message('FAIL', 'COD: 3')

    try:
        if len(definit) > 0:
            next_modules(following=module4)
            module4.table.setRowCount(len(definit))
            btns = definit.copy()
            chkBoxItem = definit.copy()
            for i in range(len(btns)):
                btns[i] = QPushButton(module4.table)
                btns[i].setText('VIEW')
                chkBoxItem[i] = QTableWidgetItem()
                chkBoxItem[i].setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
                chkBoxItem[i].setCheckState(QtCore.Qt.Unchecked)
            for row in range(len(definit)):
                module4.table.setItem(row, 0, QTableWidgetItem(projectname[definit[row]]))
                module4.table.setItem(row, 1, QTableWidgetItem(year[definit[row]]))
                module4.table.setItem(row, 2, QTableWidgetItem(turbina[definit[row]]))
                module4.table.setCellWidget(row, 3, btns[row])
                btns[row].clicked.connect(partial(show_image, definit[row]))
                module4.table.setItem(row, 4, chkBoxItem[row])
                row += row

        else:
            show_message('Warning', "This structural verification, isn't present on this record, the calculations"
                                    " must be done, please let it know to the person in charge of the calculations.")
    except:
        show_message('FAIL', 'COD: 4')


def show_image(index):
    module5.datasheett.clear()
    n = projectname[index]
    n.replace(' ', '')
    image_name = 'modules/ranges/' + n + '.png'
    excel2img.export_img("workbook/verifiche_strutturali.xlsx", image_name, year[index], ranges[index])
    im = Image.open(image_name)
    width, height = im.size
    datasheet = QPixmap(image_name)
    datasheet = datasheet.scaled(width, height, QtCore.Qt.KeepAspectRatio)
    module5.datasheett.setPixmap(datasheet)
    module5.datasheett.resize(datasheet.width(), datasheet.height())
    module5.Layout.addWidget(module5.scrollArea)
    module5.show()


def nsearch():
    module5.close()
    del definit[:]
    module4.close()
    module1.show()


# Load all data in .xslx file
def load_paramters():
    for worksheet in wb.worksheets:
        if worksheet.title.startswith('2'):
            ws = wb[worksheet.title]
            for row in ws.rows:
                for cell in row:
                    if cell.value is not None:
                        if cell.value == 'TIPOLOGIA':
                            tipologia.append(str(ws.cell(row=cell.row + 1, column=cell.column).value))
                            year.append(worksheet.title)
                        if cell.value == 'INSTALLATION:':
                            installation.append(str(ws.cell(row=cell.row, column=cell.column + 3).value))
                        if cell.value == 'EXPOSITION:':
                            exposition.append(str(ws.cell(row=cell.row, column=cell.column + 3).value))
                        if cell.value == 'TURBINA':
                            turbina.append(str(ws.cell(row=cell.row + 1, column=cell.column).value))
                        if cell.value == 'CONFORMAZIONE STRUTTURALE':
                            structure.append(str(ws.cell(row=cell.row + 1, column=cell.column).value))
                        if cell.value == 'COMMESSA':
                            projectname.append(str(ws.cell(row=cell.row + 1, column=cell.column).value))
                        if cell.value == 'SISMA APPLICATO':
                            seismicloads.append(str(ws.cell(row=cell.row + 1, column=cell.column).value))
                        if cell.value == 'VERIFICA BLAST':
                            blastloads.append(ws.cell(row=cell.row + 1, column=cell.column).value)
                        if cell.value == 'APPLIED LOADS:':
                            thermalloads.append(str(ws.cell(row=cell.row + 2, column=cell.column).value))
                        if cell.value == 'TECHNICAL DATA ':
                            snowloads.append(str(ws.cell(row=cell.row + 7, column=cell.column).value))
                        if cell.value == 'COND. DI TRASPORTO APPLICATO':
                            transport.append(str(ws.cell(row=cell.row + 1, column=cell.column).value))
                        if cell.value == 'VENTO APPLICATO':
                            windloads.append(ws.cell(row=cell.row + 1, column=cell.column).value)
                        if cell.value == 'RANGE':
                            ranges.append(ws.cell(row=cell.row + 1, column=cell.column).value)


def next_first():
    first.close()
    module1.show()


if __name__ == '__main__':
    app = QtWidgets.QApplication([])

    try:
        # Create instances of ui files
        module1 = uic.loadUi('modules/module1.ui')
        module2 = uic.loadUi('modules/module2.ui')
        module3 = uic.loadUi('modules/module3.ui')
        module4 = uic.loadUi('modules/module4.ui')
        module5 = uic.loadUi('modules/module5.ui')
        first = uic.loadUi('modules/wload.ui')
        first.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        # Set names of the forms
        module1.setWindowTitle("Structure Search Engine")
        module2.setWindowTitle("Structure Search Engine")
        module3.setWindowTitle("Structure Search Engine")
        module4.setWindowTitle("Structure Search Engine")
        module5.setWindowTitle("Structure Search Engine")
        first.setWindowIcon(QtGui.QIcon('modules/images/landing.png'))
        module1.setWindowIcon(QtGui.QIcon('modules/images/landing.png'))
        module2.setWindowIcon(QtGui.QIcon('modules/images/landing.png'))
        module3.setWindowIcon(QtGui.QIcon('modules/images/landing.png'))
        module4.setWindowIcon(QtGui.QIcon('modules/images/landing.png'))
        module5.setWindowIcon(QtGui.QIcon('modules/images/landing.png'))
    except:
        show_message('Interface Error', 'An error has been detected, please contact the person in charge. (5)')
        quit()

    # Set images an logos on a variable
    imagen = QPixmap('modules/images/load.png')
    logo = QPixmap('modules/images/sistemi.png')
    warning = QPixmap('modules/images/warning.png')
    # Assign this variables to the labels
    first.imagen.setPixmap(imagen)
    module1.logo.setPixmap(logo)
    module1.warning.setPixmap(warning)
    module2.logo.setPixmap(logo)
    module2.warning.setPixmap(warning)
    module3.logo.setPixmap(logo)
    module3.warning.setPixmap(warning)
    module4.logo.setPixmap(logo)
    module4.warning.setPixmap(warning)
    # ruta
    route = 'workbook/verifiche_strutturali.xlsx'
    try:
        pass
        wb = load_workbook(filename=route, read_only=True)
        load_paramters()
    except:
        file = route.split('/')
        show_message('File Error', 'The name of the file should be "' + file[1] + '", inside the folder \n"' + file[0] +
                     '". If this is OK and this error continues, please contact the person in charge. (4)')
        quit()

    # Configs the combo box
    combobox_turbine = list(set(turbina))
    combobox_turbine.sort()
    module1.turbine.addItems(combobox_turbine)
    module1.turbine_2.addItems(combobox_turbine)

    # Events in all modules
    # Events in moudule1
    module1.all.clicked.connect(see_all)
    module1.exit.clicked.connect(out)
    module1.concerning.clicked.connect(about)
    # Tab 1
    module1.process.clicked.connect(next_advanced)
    # Tab 2
    module1.basic.clicked.connect(next_basic)
    # Events in module2
    module2.next.clicked.connect(on_next)
    module2.exit.clicked.connect(out)
    module2.back.clicked.connect(back)
    # Events in module3
    module3.next.clicked.connect(on_next)
    module3.exit.clicked.connect(out)
    module3.back.clicked.connect(back)
    # Events in module4
    module4.exit.clicked.connect(out)
    module4.back.clicked.connect(back)
    module4.nsearch.clicked.connect(nsearch)
    # event in first
    first.next_btn.clicked.connect(next_first)
    # Exec app
    first.show()
    app.exec()